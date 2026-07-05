"""
视图混入类（View Mixins）— 组合到类视图中的可复用功能。

知识点:
  1. Mixin: 提供特定功能的小型类，通过多继承组合
  2. MRO (Method Resolution Order): `ClassName.__mro__` 可查看
  3. `super()` 在 Mixin 中的协作式调用
  4. `__init_subclass__` — 子类初始化钩子
"""
import logging
from typing import Any, Optional

from django.db.models import QuerySet
from rest_framework import status
from rest_framework.response import Response

from common.response import ApiResponse, ok

logger = logging.getLogger(__name__)


class ActionLogMixin:
    """
    操作日志混入类 — 自动记录增删改查操作。

    知识点:
      - Mixin 类通常不继承任何基类（或只继承 object）
      - 覆盖 DRF 视图的方法（如 perform_create）而不影响父类
    """

    action_log_fields: tuple[str, ...] = ()

    def perform_create(self, serializer) -> object:
        """创建后记录日志。"""
        instance = serializer.save()
        self._log_action("CREATE", instance)
        return instance

    def perform_update(self, serializer) -> object:
        """更新后记录日志。"""
        instance = serializer.save()
        self._log_action("UPDATE", instance)
        return instance

    def perform_destroy(self, instance) -> None:
        """删除后记录日志。"""
        self._log_action("DELETE", instance)
        instance.delete()

    def _log_action(self, action: str, instance: Any) -> None:
        """
        记录操作日志。

        知识点:
          - hasattr: 安全检查属性是否存在
          - 鸭子类型: 关注对象行为而非类型
        """
        try:
            # 延迟导入避免循环依赖
            from apps.logs.models import OperationLog

            # 知识点: 海象运算符 :=
            if user := getattr(self.request, "user", None):
                if user.is_authenticated:
                    OperationLog.objects.create(
                        user=user,
                        action=action,
                        model_name=instance.__class__.__name__,
                        object_id=str(instance.pk),
                        detail=str(instance),
                        ip_address=self._get_client_ip(),
                    )
        except Exception as e:
            # 知识点: 日志记录不应影响主流程
            logger.warning(f"操作日志记录失败: {e}", exc_info=True)

    def _get_client_ip(self) -> str:
        """获取客户端 IP。"""
        x_forwarded_for = self.request.META.get("HTTP_X_FORWARDED_FOR")
        if x_forwarded_for:
            return x_forwarded_for.split(",")[0].strip()
        return self.request.META.get("REMOTE_ADDR", "")


class CachedListMixin:
    """
    列表缓存混入类 — 为列表接口添加缓存。

    知识点:
      - 装饰器模式在 Mixin 中的应用
      - cache.get / cache.set 使用
      - 超时时间可配置
    """

    cache_timeout: int = 300  # 5 分钟
    cache_key_prefix: str = "list"

    def get_cache_key(self) -> str:
        """生成缓存键。"""
        return f"{self.cache_key_prefix}:{self.request.get_full_path()}"

    def list(self, request, *args, **kwargs) -> Response:
        """
        带缓存的列表查询。

        知识点:
          - from django.core.cache import cache
          - 缓存命中则直接返回，未命中则查询并设置缓存
        """
        from django.core.cache import cache

        cache_key = self.get_cache_key()
        cached_data = cache.get(cache_key)

        if cached_data is not None:
            return Response(cached_data)

        response = super().list(request, *args, **kwargs)

        # 只缓存成功响应
        if response.status_code == status.HTTP_200_OK:
            cache.set(cache_key, response.data, self.cache_timeout)

        return response


class ExportMixin:
    """
    数据导出混入类 — 为视图添加导出功能。

    知识点:
      - 生成器 (yield): 逐行生成数据，内存友好
      - yield from: 委托给子生成器（Python 3.3+）
      - openpyxl: Excel 操作库
    """

    def export_data(self, queryset: QuerySet, fields: list[str]) -> bytes:
        """
        导出数据为 Excel 字节流。

        知识点:
          - 延迟导入: 只在需要时加载（节省启动时间）
          - with 上下文管理器: 自动关闭文件
          - BytesIO: 内存中的字节缓冲区
        """
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "导出数据"

        # ─── 写入表头 ────────────────────────────────────────
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        for col, field in enumerate(fields, 1):
            cell = ws.cell(row=1, column=col, value=field)
            cell.font = header_font
            cell.fill = header_fill

        # ─── 写入数据（逐行 yield） ──────────────────────────
        # 知识点: enumerate 从 1 开始，生成器逐行产出
        for row_idx, obj in enumerate(queryset.iterator(), 2):
            for col_idx, field in enumerate(fields, 1):
                value = getattr(obj, field, "")
                if callable(value):
                    value = value()
                if not isinstance(value, (str, int, float, bool, type(None))):
                    value = str(value)
                ws.cell(row=row_idx, column=col_idx, value=value)

        # 知识点: 海象运算符 + 上下文管理器
        if (buffer := BytesIO()) is not None:
            wb.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()

        return b""


class CacheInvalidationMixin:
    """
    缓存失效混入类 — 写操作后自动删除相关缓存。

    知识点:
      - 与 CachedListMixin 配合使用
      - cache.delete_pattern (如 django-redis 支持)
      - cache.delete_many
    """

    cache_keys_to_invalidate: list[str] = []

    def perform_create(self, serializer) -> object:
        instance = super().perform_create(serializer)
        self._invalidate_cache()
        return instance

    def perform_update(self, serializer) -> object:
        instance = super().perform_update(serializer)
        self._invalidate_cache()
        return instance

    def perform_destroy(self, instance) -> None:
        super().perform_destroy(instance)
        self._invalidate_cache()

    def _invalidate_cache(self) -> None:
        """使相关缓存失效。"""
        from django.core.cache import cache

        for key in self.cache_keys_to_invalidate:
            cache.delete(key)
            logger.debug(f"缓存已失效: {key}")
